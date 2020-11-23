from openpyxl import load_workbook
import os
import pandas as pd

def iterateThroughFilesInFolder(folderPath):
    return os.listdir(folderPath)

def getFieldsFromXlsx(filePath, fields):
    workbook = load_workbook(filePath)
    sheet1 = workbook.get_sheet_by_name("clinical")
    iter = sheet1.iter_rows()
    row0 = [cell.value for cell in next(iter)]
    fieldToIndexMap = {}
    for index, thing in enumerate(row0):
        fieldToIndexMap[thing] = index
    outPutList = []
    for row in iter:
        newDict = {}
        rowAsList = [cell.value for cell in row]
        for thing in fields:
            newDict[thing] = rowAsList[fieldToIndexMap[thing]]
        outPutList.append(newDict)
    return outPutList


def getExtractEvaluationInformation(filePath):
    workbook = load_workbook(filePath)
    sheet1 = workbook.get_sheet_by_name("sheet1")
    iter = sheet1.iter_rows()
    row0 = [cell.value for cell in next(iter)]

    indexToFieldMap = {}
    for index, thing in enumerate(row0):
        indexToFieldMap[index] = thing
    indexToFieldMap[2] = "time"

    outPutList = []
    for row in iter:
        patientReportList = []
        outPutList.append(patientReportList)
        rowAsList = list([cell.value for cell in row])
        currReportDict = None
        for index,thing in enumerate(rowAsList):
            if(index == 0 or thing == None):
                continue
            if "post-MR" in indexToFieldMap[index]:
                if currReportDict != None:
                    if "cancer status" not in currReportDict and "evidence of cancer" not in currReportDict:
                        patientReportList = patientReportList[:-1]
                currReportDict = {}
                patientReportList.append(currReportDict)
                currReportDict["post-MR"] = str(thing).strip(" ").split(',')
            else:

                if thing != "" and "#" not in str(thing):
                    currReportDict[indexToFieldMap[index]] = str(thing)
        if currReportDict != None:
            if "cancer status" not in currReportDict and "evidence of cancer" not in currReportDict:
                patientReportList = patientReportList[:-1]
    return outPutList


def get_report_labels():
    """
        Input: None
        Output: dictionary of {report ID : evidence of cancer} from 'evaluation of reports.xlsx'
    """
    data = pd.read_excel("../data/evaluation of reports.xlsx")
    report_labels = {}

    # loop through each column in the excel sheet of report ids - evidence of cancer
    for i in range(1, 21):
        report_col = "post-MR "
        report_col += str(i)
        evidence_col = "evidence of cancer" 
        if i != 1:
            # pandas adds a ".number" to each repeated column name
            evidence_col += "."
            evidence_col += str(i - 1)
        reports = data[report_col]
        labels = data[evidence_col]
        temp = dict(zip(reports, labels))
        report_labels.update(temp)
    return report_labels
    
def get_report_impressions():
    """
        Input: None
         Output: dictionary of {report ids : impressions} from /reports directory
    """
    for filename in os.scandir("../data/reports"):
        # grab fourth to last token which should pertain to report ID
        name = filename.name
        tokenized_name = re.split('_|-', name)
        report_id = tokenized_name[-4]
        report = open("../data/reports/" + name, "r").readline()
        pass