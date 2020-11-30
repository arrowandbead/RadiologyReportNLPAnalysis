from openpyxl import load_workbook
import os
import pandas as pd
import re
import math

def iterateThroughFilesInFolder(folderPath):
    return os.listdir(folderPath)

def getFieldsFromXlsx(filePath, fields):
    workbook = load_workbook(filePath)
    sheet1 = workbook.get_sheet_by_name("clinical")
    iter = sheet1.iter_rows()
    row0 = [cell.value for cell in next(iter)]
    fieldToIndexMap = {}
    for index, thing in enumerate(row0):
        fieldToIndexMap[thing] = index
    outPutList = []
    for row in iter:
        newDict = {}
        rowAsList = [cell.value for cell in row]
        for thing in fields:
            newDict[thing] = rowAsList[fieldToIndexMap[thing]]
        outPutList.append(newDict)
    return outPutList


def getExtractEvaluationInformation(filePath):
    workbook = load_workbook(filePath)
    sheet1 = workbook.get_sheet_by_name("sheet1")
    iter = sheet1.iter_rows()
    row0 = [cell.value for cell in next(iter)]

    indexToFieldMap = {}
    for index, thing in enumerate(row0):
        indexToFieldMap[index] = thing
    indexToFieldMap[2] = "time"

    outPutList = []
    for row in iter:
        patientReportList = []
        outPutList.append(patientReportList)
        rowAsList = list([cell.value for cell in row])
        currReportDict = None
        for index,thing in enumerate(rowAsList):
            if(index == 0 or thing == None):
                continue
            if "post-MR" in indexToFieldMap[index]:
                if currReportDict != None:
                    if "cancer status" not in currReportDict and "evidence of cancer" not in currReportDict:
                        patientReportList = patientReportList[:-1]
                currReportDict = {}
                patientReportList.append(currReportDict)
                currReportDict["post-MR"] = str(thing).strip(" ").split(',')
            else:

                if thing != "" and "#" not in str(thing):
                    currReportDict[indexToFieldMap[index]] = str(thing)
        if currReportDict != None:
            if "cancer status" not in currReportDict and "evidence of cancer" not in currReportDict:
                patientReportList = patientReportList[:-1]
    return outPutList


def get_report_labels():
    """
        Input: None
        Output: dictionary of {report ID : evidence of cancer} from 'evaluation of reports.xlsx'
    """
    data = pd.read_excel("../data/evaluation of reports.xlsx")
    report_labels = {}

    # loop through each column in the excel sheet of report ids - evidence of cancer
    for i in range(1, 21):
        report_col = "post-MR "
        report_col += str(i)
        evidence_col = "evidence of cancer" 
        if i != 1:
            # pandas adds a ".number" to each repeated column name
            evidence_col += "."
            evidence_col += str(i - 1)
        reports = data[report_col].astype(str)
        labels = data[evidence_col]
        temp = dict(zip(reports, labels))
        report_labels.update(temp)
    return report_labels
    
def get_report_impressions():
    """
        Input: None
         Output: dictionary of {report ids : impressions} from /reports directory
    """
    report_impressions = {}
    for filename in os.scandir("../data/reports"):
        name = filename.name
        if (name == ".DS_Store"):
            continue
        # print("name of file:", name)
        tokenized_name = re.split('_|-', name)
        # print("tok name", tokenized_name)
        report_id = tokenized_name[-4]
        if report_id == "6320734":
            print(open("../data/reports/" + name, "r").read())
        # print(list(tokenized_name[-1])[-3:])
        if (list(tokenized_name[-1])[-3:] != ['t','x','t']):
            continue
        # print("--------------------------------------")
        # print(open("../data/reports/" + name, "r").read())
        report = open("../data/reports/" + name, "r").readlines()
        add_to_impression = False
        impression = ""
        for line in report:
            tokenize_line = line.split()
            if (len(tokenize_line) == 0):
                # blank line
                continue
            first_word = tokenize_line[0].strip()
            if first_word == "END":
                # end of impression
                add_to_impression = False
            if add_to_impression:
                impression += line.strip() + " "
            if first_word == "IMPRESSION" or first_word == "IMPRESSION:":
                impression = ""
                first_line = tokenize_line[1:]
                for word in first_line:
                    impression += word + " "
                add_to_impression = True
        # print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        # print("Impression extracted:")
        # print(impression)
        # print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        report_impressions.update({report_id:impression})
    return report_impressions

def remove_nan_labels():
    label_dictionary = get_report_labels()
    for key in label_dictionary:
        if math.isnan(key):
            del label_dictionary[key]


# OPTIONAL TODO: Remove list numbers from impressions in numbered lists
# OPTIONAL TODO: Even formatting in report ids (remove spaces, \n's, etc.)